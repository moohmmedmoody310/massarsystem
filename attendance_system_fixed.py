from flask import Flask, request, jsonify, session, send_file
from datetime import datetime, timedelta
import sqlite3
import json
from database import get_db, close_db
from pdf_generator_advanced import generate_ai_student_report
import os

class AdvancedAttendanceSystem:
    def __init__(self):
        self.attendance_types = {
            'present': {'name': 'حاضر', 'color': '#27ae60', 'points': 3},
            'absent': {'name': 'غائب', 'color': '#e74c3c', 'points': 0},
            'late': {'name': 'متأخر', 'color': '#f39c12', 'points': 2},
            'excused': {'name': 'معذور', 'color': '#3498db', 'points': 2},
            'sick': {'name': 'مريض', 'color': '#9b59b6', 'points': 1}
        }
    
    def get_today_students(self, group_id=None):
        """الحصول على جميع الطلاب ليوم اليوم"""
        db = get_db()
        cursor = db.cursor()
        
        query = '''
            SELECT s.id, s.name, s.phone, s.parent_name, s.parent_phone, 
                   s.group_id, g.name as group_name, g.subject_id, sub.name as subject_name
            FROM students s
            LEFT JOIN groups g ON s.group_id = g.id
            LEFT JOIN subjects sub ON g.subject_id = sub.id
            WHERE s.status = 'active'
        '''
        
        params = []
        if group_id:
            query += ' AND s.group_id = ?'
            params.append(group_id)
        
        query += ' ORDER BY g.name, s.name'
        
        cursor.execute(query, params)
        students = cursor.fetchall()
        db.close()
        
        return [dict(student) for student in students]
    
    def mark_attendance_batch(self, attendance_data):
        """تسجيل الحضور والغياب لجميع الطلاب دفعة واحدة"""
        db = get_db()
        cursor = db.cursor()
        
        today = datetime.now().strftime('%Y-%m-%d')
        success_count = 0
        errors = []
        
        for student_id, status in attendance_data.items():
            try:
                # التحقق من وجود سجل سابق لليوم
                cursor.execute('''
                    SELECT id FROM attendance 
                    WHERE student_id = ? AND date = ?
                ''', (student_id, today))
                
                existing = cursor.fetchone()
                
                if existing:
                    # تحديث السجل الموجود
                    cursor.execute('''
                        UPDATE attendance 
                        SET status = ?, notes = ?, updated_at = ?
                        WHERE id = ?
                    ''', (status['status'], status.get('notes', ''), datetime.now(), existing[0]))
                else:
                    # إنشاء سجل جديد
                    cursor.execute('''
                        INSERT INTO attendance (student_id, group_id, date, status, notes, created_at)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (student_id, status.get('group_id'), today, status['status'], 
                          status.get('notes', ''), datetime.now()))
                
                success_count += 1
                
            except Exception as e:
                errors.append(f"خطأ في تسجيل الطالب {student_id}: {str(e)}")
        
        db.commit()
        db.close()
        
        return {
            'success': True,
            'marked_count': success_count,
            'errors': errors,
            'date': today
        }
    
    def get_attendance_report(self, date_from=None, date_to=None, student_id=None, group_id=None):
        """الحصول على تقرير الحضور والغياب"""
        db = get_db()
        cursor = db.cursor()
        
        query = '''
            SELECT a.*, s.name as student_name, s.phone, s.parent_name, s.parent_phone,
                   g.name as group_name, sub.name as subject_name
            FROM attendance a
            JOIN students s ON a.student_id = s.id
            LEFT JOIN groups g ON a.group_id = g.id
            LEFT JOIN subjects sub ON g.subject_id = sub.id
            WHERE 1=1
        '''
        
        params = []
        
        if date_from:
            query += ' AND a.date >= ?'
            params.append(date_from)
        
        if date_to:
            query += ' AND a.date <= ?'
            params.append(date_to)
        
        if student_id:
            query += ' AND a.student_id = ?'
            params.append(student_id)
        
        if group_id:
            query += ' AND a.group_id = ?'
            params.append(group_id)
        
        query += ' ORDER BY a.date DESC, s.name'
        
        cursor.execute(query, params)
        records = cursor.fetchall()
        db.close()
        
        return [dict(record) for record in records]
    
    def get_student_attendance_stats(self, student_id, days=30):
        """الحصول على إحصائيات الحضور للطالب"""
        db = get_db()
        cursor = db.cursor()
        
        # الحصول على سجل الحضور
        cursor.execute('''
            SELECT status, date, notes
            FROM attendance
            WHERE student_id = ? AND date >= date('now', '-{} days')
            ORDER BY date DESC
        '''.format(days), (student_id,))
        
        records = cursor.fetchall()
        
        # حساب الإحصائيات
        stats = {
            'total_days': len(records),
            'present': 0,
            'absent': 0,
            'late': 0,
            'excused': 0,
            'sick': 0,
            'attendance_rate': 0,
            'recent_trend': [],
            'points': 0
        }
        
        for record in records:
            status = record['status']
            stats[status] = stats.get(status, 0) + 1
            stats['points'] += self.attendance_types.get(status, {}).get('points', 0)
            
            # تحليل الاتجاه الأخير
            stats['recent_trend'].append({
                'date': record['date'],
                'status': status,
                'points': self.attendance_types.get(status, {}).get('points', 0)
            })
        
        # حساب نسبة الحضور
        if stats['total_days'] > 0:
            stats['attendance_rate'] = (stats['present'] / stats['total_days']) * 100
        
        db.close()
        return stats
    
    def generate_daily_attendance_sheet(self, group_id=None):
        """إنشاء ورقة حضور يومية"""
        students = self.get_today_students(group_id)
        today = datetime.now().strftime('%Y-%m-%d')
        
        # التحقق من وجود سجلات سابقة لليوم
        db = get_db()
        cursor = db.cursor()
        
        existing_records = {}
        if students:
            student_ids = [str(s['id']) for s in students]
            placeholders = ','.join(['?'] * len(student_ids))
            cursor.execute(f'''
                SELECT student_id, status, notes
                FROM attendance
                WHERE date = ? AND student_id IN ({placeholders})
            ''', [today] + student_ids)
            
            for record in cursor.fetchall():
                existing_records[str(record['student_id'])] = {
                    'status': record['status'],
                    'notes': record['notes']
                }
        
        db.close()
        
        # تجهيز بيانات الورقة
        attendance_sheet = {
            'date': today,
            'group_id': group_id,
            'students': []
        }
        
        for student in students:
            student_data = {
                'id': student['id'],
                'name': student['name'],
                'group_name': student['group_name'],
                'subject_name': student['subject_name'],
                'phone': student['phone'],
                'parent_name': student['parent_name'],
                'parent_phone': student['parent_phone'],
                'status': existing_records.get(str(student['id']), {}).get('status', 'present'),
                'notes': existing_records.get(str(student['id']), {}).get('notes', ''),
                'attendance_types': self.attendance_types
            }
            attendance_sheet['students'].append(student_data)
        
        return attendance_sheet
    
    def send_attendance_notifications(self, attendance_data):
        """إرسال إشعارات الحضور لأولياء الأمور"""
        notifications_sent = []
        
        for student_id, data in attendance_data.items():
            if data['status'] in ['absent', 'late']:
                # إرسال إشعار لولي الأمر
                message = self.create_attendance_message(data)
                # هنا يمكن إضافة كود إرسال الرسالة عبر WhatsApp أو SMS
                notifications_sent.append({
                    'student_id': student_id,
                    'parent_phone': data.get('parent_phone'),
                    'message': message,
                    'status': 'sent'
                })
        
        return notifications_sent
    
    def create_attendance_message(self, data):
        """إنشاء رسالة حضور لولي الأمر"""
        status_text = self.attendance_types.get(data['status'], {}).get('name', 'غير محدد')
        date = datetime.now().strftime('%Y/%m/%d')
        
        message = f"""
        مركز مسار التعليمي
        
        السلام عليكم ورحمة الله وبركاته،
        
        نود إعلامكم بأن ابنكم/ابنتكم {data['student_name']} 
        اليوم {date}
        الحالة: {status_text}
        
        مع أطيب التحيات،
        مركز مسار التعليمي
        """
        
        return message.strip()
    
    def export_attendance_excel(self, date_from, date_to, group_id=None):
        """تصدير بيانات الحضور إلى Excel"""
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
        
        # الحصول على البيانات
        records = self.get_attendance_report(date_from, date_to, group_id=group_id)
        
        # إنشاء ملف Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير الحضور والغياب"
        
        # إعدادات التنسيق
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1a2615', end_color='1a2615', fill_type='solid')
        cell_alignment = Alignment(horizontal='center', vertical='center')
        
        # العناوين
        headers = ['التاريخ', 'اسم الطالب', 'المجموعة', 'المادة', 'الحالة', 'ملاحظات']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = cell_alignment
        
        # البيانات
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record['date'])
            ws.cell(row=row, column=2, value=record['student_name'])
            ws.cell(row=row, column=3, value=record['group_name'] or '-')
            ws.cell(row=row, column=4, value=record['subject_name'] or '-')
            
            # تلوين الحالة
            status_cell = ws.cell(row=row, column=5, value=self.attendance_types.get(record['status'], {}).get('name', record['status']))
            status_color = self.attendance_types.get(record['status'], {}).get('color', 'FFFFFF')
            status_fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
            status_cell.fill = status_fill
            status_cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            status_cell.alignment = cell_alignment
            
            ws.cell(row=row, column=6, value=record.get('notes', '-'))
        
        # ضبط عرض الأعمدة
        column_widths = [12, 20, 15, 15, 10, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # حفظ الملف
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

# إنشاء كائن نظام الحضور
attendance_system = AdvancedAttendanceSystem()
