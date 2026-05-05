from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from database import get_db, init_db
from pdf_generator import generate_student_receipt, generate_monthly_report
from pdf_generator_advanced import generate_ai_student_report, generate_advanced_student_receipt, generate_financial_report
from attendance_system import attendance_system
import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
import hashlib
import os
import io
from datetime import datetime, date
from functools import wraps

app = Flask(__name__)
app.secret_key = 'gamal_naser_center_2024_secret_key'

# إضافة CORS headers
@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

# إنشاء قاعدة البيانات عند بدء التشغيل
init_db()

MONTHS_AR = {
    1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
    5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
    9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
}

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

# ==================== PAGES ====================

@app.route('/')
@login_required
def index():
    return render_template('index.html', user=session.get('user_name'))

@app.route('/login')
def login_page():
    if 'user_id' in session:
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

# ==================== AUTH API ====================

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username', '')
    password = hashlib.sha256(data.get('password', '').encode()).hexdigest()
    
    db = get_db()
    user = db.execute(
        "SELECT * FROM users WHERE username=? AND password=?", (username, password)
    ).fetchone()
    db.close()
    
    if user:
        session['user_id'] = user['id']
        session['user_name'] = user['full_name'] or user['username']
        session['user_role'] = user['role']
        return jsonify({'success': True, 'name': session['user_name'], 'role': user['role']})
    return jsonify({'success': False, 'message': 'اسم المستخدم أو كلمة المرور غلط'})

# ==================== DASHBOARD ====================

@app.route('/api/dashboard')
@login_required
def dashboard():
    db = get_db()
    now = datetime.now()
    month = now.month
    year = now.year
    
    total_students = db.execute("SELECT COUNT(*) FROM students WHERE status='active'").fetchone()[0]
    total_teachers = db.execute("SELECT COUNT(*) FROM teachers WHERE status='active'").fetchone()[0]
    total_groups = db.execute("SELECT COUNT(*) FROM groups WHERE status='active'").fetchone()[0]
    
    monthly_income = db.execute(
        "SELECT COALESCE(SUM(amount),0) FROM payments WHERE month=? AND year=?",
        (month, year)
    ).fetchone()[0]
    
    monthly_expenses = db.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE strftime('%m',date)=? AND strftime('%Y',date)=?",
        (f'{month:02d}', str(year))
    ).fetchone()[0]
    
    # الطلاب لم يدفعوا هذا الشهر
    paid_students = db.execute(
        "SELECT COUNT(DISTINCT student_id) FROM payments WHERE month=? AND year=?",
        (month, year)
    ).fetchone()[0]
    unpaid_students = total_students - paid_students
    
    # آخر 5 دفعات
    recent_payments = db.execute('''
        SELECT p.*, s.name as student_name
        FROM payments p JOIN students s ON p.student_id = s.id
        ORDER BY p.payment_date DESC LIMIT 5
    ''').fetchall()
    
    # إحصائيات الشهور الستة الأخيرة
    monthly_stats = []
    for i in range(5, -1, -1):
        m = month - i
        y = year
        while m <= 0:
            m += 12
            y -= 1
        income = db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM payments WHERE month=? AND year=?",
            (m, y)
        ).fetchone()[0]
        monthly_stats.append({
            'month': MONTHS_AR.get(m, str(m)),
            'income': income
        })
    
    db.close()
    
    return jsonify({
        'total_students': total_students,
        'total_teachers': total_teachers,
        'total_groups': total_groups,
        'monthly_income': monthly_income,
        'monthly_expenses': monthly_expenses,
        'net_income': monthly_income - monthly_expenses,
        'unpaid_students': unpaid_students,
        'paid_students': paid_students,
        'recent_payments': [dict(p) for p in recent_payments],
        'monthly_stats': monthly_stats,
        'current_month': MONTHS_AR.get(month, ''),
        'current_year': year
    })

# ==================== STUDENTS API ====================

@app.route('/api/students', methods=['GET'])
@login_required
def get_students():
    db = get_db()
    search = request.args.get('search', '')
    grade_id = request.args.get('grade_id', '')
    group_id = request.args.get('group_id', '')
    status = request.args.get('status', 'active')
    
    query = '''
        SELECT s.*, g.name as grade_name, gr.name as group_name,
               t.name as teacher_name
        FROM students s
        LEFT JOIN grades g ON s.grade_id = g.id
        LEFT JOIN groups gr ON s.group_id = gr.id
        LEFT JOIN teachers t ON gr.teacher_id = t.id
        WHERE 1=1
    '''
    params = []
    
    if status:
        query += " AND s.status=?"
        params.append(status)
    if search:
        query += " AND (s.name LIKE ? OR s.phone LIKE ? OR s.parent_phone LIKE ?)"
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])
    if grade_id:
        query += " AND s.grade_id=?"
        params.append(grade_id)
    if group_id:
        query += " AND s.group_id=?"
        params.append(group_id)
    
    query += " ORDER BY s.name"
    students = db.execute(query, params).fetchall()
    db.close()
    return jsonify([dict(s) for s in students])

@app.route('/api/students', methods=['POST'])
@login_required
def add_student():
    data = request.json
    db = get_db()
    try:
        db.execute('''
            INSERT INTO students (name, phone, parent_name, parent_phone, parent_whatsapp,
                                  grade_id, group_id, national_id, address, birth_date, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            data.get('name'), data.get('phone'), data.get('parent_name'),
            data.get('parent_phone'), data.get('parent_whatsapp'),
            data.get('grade_id'), data.get('group_id'), data.get('national_id'),
            data.get('address'), data.get('birth_date'), data.get('notes')
        ))
        db.commit()
        return jsonify({'success': True, 'message': 'تم إضافة الطالب بنجاح'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        db.close()

@app.route('/api/students/<int:sid>', methods=['GET'])
@login_required
def get_student(sid):
    db = get_db()
    student = db.execute('''
        SELECT s.*, g.name as grade_name, gr.name as group_name
        FROM students s
        LEFT JOIN grades g ON s.grade_id = g.id
        LEFT JOIN groups gr ON s.group_id = gr.id
        WHERE s.id=?
    ''', (sid,)).fetchone()
    
    payments = db.execute('''
        SELECT * FROM payments WHERE student_id=? ORDER BY year DESC, month DESC
    ''', (sid,)).fetchall()
    
    attendance = db.execute('''
        SELECT * FROM attendance WHERE student_id=? ORDER BY date DESC LIMIT 30
    ''', (sid,)).fetchall()
    
    results = db.execute('''
        SELECT gr.*, s.name as subject_name
        FROM grades_results gr
        LEFT JOIN subjects s ON gr.subject_id = s.id
        WHERE gr.student_id=? ORDER BY gr.exam_date DESC
    ''', (sid,)).fetchall()
    
    db.close()
    
    if not student:
        return jsonify({'error': 'الطالب غير موجود'}), 404
    
    return jsonify({
        'student': dict(student),
        'payments': [dict(p) for p in payments],
        'attendance': [dict(a) for a in attendance],
        'results': [dict(r) for r in results]
    })

@app.route('/api/students/<int:sid>', methods=['PUT'])
@login_required
def update_student(sid):
    data = request.json
    db = get_db()
    try:
        db.execute('''
            UPDATE students SET name=?, phone=?, parent_name=?, parent_phone=?,
            parent_whatsapp=?, grade_id=?, group_id=?, national_id=?,
            address=?, birth_date=?, notes=?, status=?
            WHERE id=?
        ''', (
            data.get('name'), data.get('phone'), data.get('parent_name'),
            data.get('parent_phone'), data.get('parent_whatsapp'),
            data.get('grade_id'), data.get('group_id'), data.get('national_id'),
            data.get('address'), data.get('birth_date'), data.get('notes'),
            data.get('status', 'active'), sid
        ))
        db.commit()
        return jsonify({'success': True, 'message': 'تم تحديث بيانات الطالب'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        db.close()

@app.route('/api/students/<int:sid>', methods=['DELETE'])
@login_required
def delete_student(sid):
    db = get_db()
    db.execute("UPDATE students SET status='inactive' WHERE id=?", (sid,))
    db.commit()
    db.close()
    return jsonify({'success': True, 'message': 'تم حذف الطالب'})

# ==================== PAYMENTS API ====================

@app.route('/api/payments', methods=['POST'])
@login_required
def add_payment():
    data = request.json
    db = get_db()
    try:
        existing = db.execute(
            "SELECT id FROM payments WHERE student_id=? AND month=? AND year=?",
            (data.get('student_id'), data.get('month'), data.get('year'))
        ).fetchone()
        
        if existing:
            return jsonify({'success': False, 'message': 'الطالب دفع هذا الشهر بالفعل'})
        
        cursor = db.execute('''
            INSERT INTO payments (student_id, amount, month, year, payment_method, notes, received_by)
            VALUES (?,?,?,?,?,?,?)
        ''', (
            data.get('student_id'), data.get('amount'), data.get('month'),
            data.get('year'), data.get('payment_method', 'cash'),
            data.get('notes'), session.get('user_name')
        ))
        payment_id = cursor.lastrowid
        db.commit()
        
        return jsonify({'success': True, 'message': 'تم تسجيل الدفعة بنجاح', 'payment_id': payment_id})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        db.close()

@app.route('/api/payments/<int:pid>/receipt')
@login_required
def get_receipt(pid):
    db = get_db()
    payment = db.execute("SELECT * FROM payments WHERE id=?", (pid,)).fetchone()
    if not payment:
        return jsonify({'error': 'الدفعة غير موجودة'}), 404
    
    student = db.execute('''
        SELECT s.*, g.name as grade_name
        FROM students s LEFT JOIN grades g ON s.grade_id=g.id
        WHERE s.id=?
    ''', (payment['student_id'],)).fetchone()
    db.close()
    
    pdf_bytes = generate_student_receipt(dict(student), dict(payment))
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=False,
        download_name=f'receipt_{pid}.pdf'
    )

@app.route('/api/payments/monthly-status')
@login_required
def monthly_payment_status():
    month = int(request.args.get('month', datetime.now().month))
    year = int(request.args.get('year', datetime.now().year))
    
    db = get_db()
    students = db.execute('''
        SELECT s.id, s.name, s.parent_phone, s.parent_whatsapp,
               g.name as grade_name, gr.monthly_fee,
               p.amount as paid_amount, p.id as payment_id
        FROM students s
        LEFT JOIN grades g ON s.grade_id = g.id
        LEFT JOIN groups gr ON s.group_id = gr.id
        LEFT JOIN payments p ON p.student_id = s.id AND p.month=? AND p.year=?
        WHERE s.status='active'
        ORDER BY s.name
    ''', (month, year)).fetchall()
    db.close()
    
    result = []
    for s in students:
        d = dict(s)
        d['paid'] = d['payment_id'] is not None
        d['required_amount'] = d['monthly_fee'] or 0
        result.append(d)
    
    return jsonify(result)

# ==================== TEACHERS API ====================

@app.route('/api/teachers', methods=['GET'])
@login_required
def get_teachers():
    db = get_db()
    teachers = db.execute('''
        SELECT t.*, COUNT(g.id) as groups_count,
               COUNT(DISTINCT s.id) as students_count
        FROM teachers t
        LEFT JOIN groups g ON g.teacher_id = t.id AND g.status='active'
        LEFT JOIN students s ON s.group_id = g.id AND s.status='active'
        WHERE t.status='active'
        GROUP BY t.id
        ORDER BY t.name
    ''').fetchall()
    db.close()
    return jsonify([dict(t) for t in teachers])

@app.route('/api/teachers', methods=['POST'])
@login_required
def add_teacher():
    data = request.json
    db = get_db()
    try:
        db.execute('''
            INSERT INTO teachers (name, phone, whatsapp, subject, salary, notes)
            VALUES (?,?,?,?,?,?)
        ''', (data.get('name'), data.get('phone'), data.get('whatsapp'),
              data.get('subject'), data.get('salary', 0), data.get('notes')))
        db.commit()
        return jsonify({'success': True, 'message': 'تم إضافة المعلم بنجاح'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        db.close()

@app.route('/api/teachers/<int:tid>', methods=['PUT'])
@login_required
def update_teacher(tid):
    data = request.json
    db = get_db()
    db.execute('''
        UPDATE teachers SET name=?, phone=?, whatsapp=?, subject=?, salary=?, notes=?
        WHERE id=?
    ''', (data.get('name'), data.get('phone'), data.get('whatsapp'),
          data.get('subject'), data.get('salary', 0), data.get('notes'), tid))
    db.commit()
    db.close()
    return jsonify({'success': True})

@app.route('/api/teachers/<int:tid>/salary', methods=['POST'])
@login_required
def pay_teacher_salary(tid):
    data = request.json
    db = get_db()
    db.execute('''
        INSERT INTO teacher_salaries (teacher_id, amount, month, year, notes)
        VALUES (?,?,?,?,?)
    ''', (tid, data.get('amount'), data.get('month'), data.get('year'), data.get('notes')))
    db.commit()
    db.close()
    return jsonify({'success': True, 'message': 'تم صرف الراتب بنجاح'})

# ==================== GROUPS API ====================

@app.route('/api/groups', methods=['GET'])
@login_required
def get_groups():
    db = get_db()
    groups = db.execute('''
        SELECT g.*, t.name as teacher_name, gr.name as grade_name,
               s.name as subject_name,
               COUNT(st.id) as students_count
        FROM groups g
        LEFT JOIN teachers t ON g.teacher_id = t.id
        LEFT JOIN grades gr ON g.grade_id = gr.id
        LEFT JOIN subjects s ON g.subject_id = s.id
        LEFT JOIN students st ON st.group_id = g.id AND st.status='active'
        WHERE g.status='active'
        GROUP BY g.id
        ORDER BY g.name
    ''').fetchall()
    db.close()
    return jsonify([dict(g) for g in groups])

@app.route('/api/groups', methods=['POST'])
@login_required
def add_group():
    data = request.json
    db = get_db()
    try:
        db.execute('''
            INSERT INTO groups (name, grade_id, teacher_id, subject_id, day_of_week,
                               time_slot, monthly_fee, max_students)
            VALUES (?,?,?,?,?,?,?,?)
        ''', (data.get('name'), data.get('grade_id'), data.get('teacher_id'),
              data.get('subject_id'), data.get('day_of_week'), data.get('time_slot'),
              data.get('monthly_fee', 0), data.get('max_students', 20)))
        db.commit()
        return jsonify({'success': True, 'message': 'تم إضافة المجموعة بنجاح'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        db.close()

# ==================== ATTENDANCE API ====================

@app.route('/api/attendance', methods=['POST'])
@login_required
def record_attendance():
    data = request.json
    db = get_db()
    records = data.get('records', [])
    today = date.today().isoformat()
    
    for r in records:
        existing = db.execute(
            "SELECT id FROM attendance WHERE student_id=? AND date=?",
            (r['student_id'], r.get('date', today))
        ).fetchone()
        
        if existing:
            db.execute(
                "UPDATE attendance SET status=? WHERE id=?",
                (r['status'], existing['id'])
            )
        else:
            db.execute(
                "INSERT INTO attendance (student_id, group_id, date, status) VALUES (?,?,?,?)",
                (r['student_id'], r.get('group_id'), r.get('date', today), r['status'])
            )
    
    db.commit()
    db.close()
    return jsonify({'success': True, 'message': f'تم تسجيل حضور {len(records)} طالب'})

@app.route('/api/attendance/group/<int:gid>')
@login_required
def get_group_attendance(gid):
    date_param = request.args.get('date', date.today().isoformat())
    db = get_db()
    
    students = db.execute('''
        SELECT s.id, s.name,
               COALESCE(a.status, 'absent') as attendance_status
        FROM students s
        LEFT JOIN attendance a ON a.student_id = s.id AND a.date=?
        WHERE s.group_id=? AND s.status='active'
        ORDER BY s.name
    ''', (date_param, gid)).fetchall()
    db.close()
    return jsonify([dict(s) for s in students])

# ==================== GRADES/RESULTS API ====================

@app.route('/api/results', methods=['POST'])
@login_required
def add_result():
    data = request.json
    db = get_db()
    db.execute('''
        INSERT INTO grades_results (student_id, subject_id, exam_name, score, max_score, exam_date, notes)
        VALUES (?,?,?,?,?,?,?)
    ''', (data.get('student_id'), data.get('subject_id'), data.get('exam_name'),
          data.get('score'), data.get('max_score', 100), data.get('exam_date'),
          data.get('notes')))
    db.commit()
    db.close()
    return jsonify({'success': True, 'message': 'تم إضافة النتيجة بنجاح'})

# ==================== EXPENSES API ====================

@app.route('/api/expenses', methods=['GET'])
@login_required
def get_expenses():
    db = get_db()
    month = request.args.get('month', '')
    year = request.args.get('year', '')
    
    query = "SELECT * FROM expenses WHERE 1=1"
    params = []
    if month and year:
        query += " AND strftime('%m',date)=? AND strftime('%Y',date)=?"
        params.extend([f'{int(month):02d}', str(year)])
    
    query += " ORDER BY date DESC"
    expenses = db.execute(query, params).fetchall()
    db.close()
    return jsonify([dict(e) for e in expenses])

@app.route('/api/expenses', methods=['POST'])
@login_required
def add_expense():
    data = request.json
    db = get_db()
    db.execute('''
        INSERT INTO expenses (description, amount, category, notes)
        VALUES (?,?,?,?)
    ''', (data.get('description'), data.get('amount'), data.get('category'), data.get('notes')))
    db.commit()
    db.close()
    return jsonify({'success': True, 'message': 'تم إضافة المصروف بنجاح'})

# ==================== EXCEL EXPORT/IMPORT ====================

@app.route('/api/export/students')
@login_required
def export_students_excel():
    db = get_db()
    students = db.execute('''
        SELECT s.name, s.phone, s.parent_name, s.parent_phone, s.parent_whatsapp,
               g.name as grade_name, gr.name as group_name, s.address,
               s.birth_date, s.join_date, s.status, s.notes
        FROM students s
        LEFT JOIN grades g ON s.grade_id = g.id
        LEFT JOIN groups gr ON s.group_id = gr.id
        ORDER BY s.name
    ''').fetchall()
    db.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'بيانات الطلاب'
    ws.sheet_view.rightToLeft = True
    
    headers = ['اسم الطالب', 'هاتف الطالب', 'اسم ولي الأمر', 'هاتف ولي الأمر',
               'واتساب ولي الأمر', 'المرحلة الدراسية', 'المجموعة', 'العنوان',
               'تاريخ الميلاد', 'تاريخ الالتحاق', 'الحالة', 'ملاحظات']
    
    header_fill = PatternFill(start_color='1a3c5e', end_color='1a3c5e', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12, name='Arial')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[chr(64+col)].width = 18
    
    for row_idx, student in enumerate(students, 2):
        for col_idx, value in enumerate(student, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='right')
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='ecf0f1', end_color='ecf0f1', fill_type='solid')
    
    ws.row_dimensions[1].height = 30
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'students_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/api/export/payments')
@login_required
def export_payments_excel():
    month = int(request.args.get('month', datetime.now().month))
    year = int(request.args.get('year', datetime.now().year))
    
    db = get_db()
    payments = db.execute('''
        SELECT s.name, g.name as grade_name, p.amount,
               p.month, p.year, p.payment_date, p.payment_method, p.received_by, p.notes
        FROM payments p
        JOIN students s ON p.student_id = s.id
        LEFT JOIN grades g ON s.grade_id = g.id
        WHERE p.month=? AND p.year=?
        ORDER BY p.payment_date DESC
    ''', (month, year)).fetchall()
    db.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'مدفوعات {MONTHS_AR.get(month,"")} {year}'
    ws.sheet_view.rightToLeft = True
    
    headers = ['اسم الطالب', 'المرحلة', 'المبلغ', 'الشهر', 'السنة',
               'تاريخ الدفع', 'طريقة الدفع', 'استلم بواسطة', 'ملاحظات']
    
    header_fill = PatternFill(start_color='27ae60', end_color='27ae60', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF', name='Arial')
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64+col)].width = 16
    
    total = 0
    for row_idx, payment in enumerate(payments, 2):
        for col_idx, value in enumerate(payment, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='right')
        total += payment[2]
    
    last_row = len(payments) + 2
    ws.cell(row=last_row, column=1, value='الإجمالي').font = Font(bold=True)
    total_cell = ws.cell(row=last_row, column=3, value=total)
    total_cell.font = Font(bold=True, color='27ae60')
    total_cell.fill = PatternFill(start_color='ecf0f1', end_color='ecf0f1', fill_type='solid')
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'payments_{month}_{year}.xlsx'
    )

@app.route('/api/import/students', methods=['POST'])
@login_required
def import_students_excel():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'لم يتم رفع ملف'})
    
    file = request.files['file']
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    
    db = get_db()
    added = 0
    errors = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row[0]:
            continue
        try:
            db.execute('''
                INSERT INTO students (name, phone, parent_name, parent_phone, parent_whatsapp, address, notes)
                VALUES (?,?,?,?,?,?,?)
            ''', (row[0], row[1] if len(row)>1 else None,
                  row[2] if len(row)>2 else None,
                  row[3] if len(row)>3 else None,
                  row[4] if len(row)>4 else None,
                  row[7] if len(row)>7 else None,
                  row[11] if len(row)>11 else None))
            added += 1
        except Exception as e:
            errors.append(f'صف {row_idx}: {str(e)}')
    
    db.commit()
    db.close()
    
    return jsonify({
        'success': True,
        'message': f'تم استيراد {added} طالب بنجاح',
        'errors': errors
    })

# ==================== LOOKUP API ====================

@app.route('/api/grades')
@login_required
def get_grades():
    db = get_db()
    grades = db.execute("SELECT * FROM grades ORDER BY id").fetchall()
    db.close()
    return jsonify([dict(g) for g in grades])

@app.route('/api/subjects')
@login_required
def get_subjects():
    db = get_db()
    subjects = db.execute("SELECT * FROM subjects ORDER BY name").fetchall()
    db.close()
    return jsonify([dict(s) for s in subjects])

@app.route('/api/reports/monthly-pdf')
@login_required
def monthly_report_pdf():
    month = int(request.args.get('month', datetime.now().month))
    year = int(request.args.get('year', datetime.now().year))
    
    db = get_db()
    students_data = db.execute('''
        SELECT s.id, s.name, g.name as grade_name,
               COALESCE(gr.monthly_fee, 0) as required_amount,
               COALESCE(p.amount, 0) as paid_amount,
               (p.id IS NOT NULL) as paid
        FROM students s
        LEFT JOIN grades g ON s.grade_id = g.id
        LEFT JOIN groups gr ON s.group_id = gr.id
        LEFT JOIN payments p ON p.student_id = s.id AND p.month=? AND p.year=?
        WHERE s.status='active'
        ORDER BY s.name
    ''', (month, year)).fetchall()
    db.close()
    
    pdf_bytes = generate_monthly_report([dict(s) for s in students_data], month, year)
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=False,
        download_name=f'monthly_report_{month}_{year}.pdf'
    )

# ==================== PUBLIC REGISTRATION API ====================

@app.route('/api/public/register', methods=['POST'])
def public_register():
    """استقبال بيانات التسجيل من الموقع العام"""
    try:
        data = request.get_json()

        if not data:
            return jsonify({'success': False, 'message': 'لا توجد بيانات'}), 400

        # استخراج البيانات من النموذج
        name = data.get('name', '').strip()
        phone = data.get('phone', '').strip()
        parent_name = data.get('parentName', '').strip()
        age = data.get('age', '').strip()
        program = data.get('program', '').strip()
        preferred_time = data.get('preferredTime', '').strip()
        notes = data.get('notes', '').strip()

        # التحقق من البيانات الأساسية
        if not name or not phone:
            return jsonify({'success': False, 'message': 'الاسم ورقم الهاتف مطلوبان'}), 400

        # حفظ البيانات في قاعدة البيانات
        db = get_db()
        cursor = db.cursor()

        # إضافة الطالب في جدول الطلاب المقدمين
        cursor.execute('''
            INSERT INTO applicants (name, phone, parent_name, age, program, preferred_time, notes, submission_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (name, phone, parent_name, age, program, preferred_time, notes, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))

        db.commit()
        db.close()

        return jsonify({
            'success': True,
            'message': 'تم التسجيل بنجاح'
        }), 200

    except Exception as e:
        print(f"Error in public_register: {e}")
        return jsonify({'success': False, 'message': 'حدث خطأ في التسجيل'}), 500

# API endpoints للطلاب المقدمين
@app.route('/api/applicants', methods=['GET'])
def get_applicants():
    if 'user_id' not in session:
        return jsonify({'error': 'غير مصرح'}), 401
    
    search = request.args.get('search', '')
    db = get_db()
    cursor = db.cursor()
    
    if search:
        cursor.execute('''
            SELECT * FROM applicants 
            WHERE name LIKE ? OR phone LIKE ? OR parent_name LIKE ?
            ORDER BY submission_date DESC
        ''', (f'%{search}%', f'%{search}%', f'%{search}%'))
    else:
        cursor.execute('SELECT * FROM applicants ORDER BY submission_date DESC')
    
    applicants = cursor.fetchall()
    db.close()
    
    return jsonify({'applicants': [dict(applicant) for applicant in applicants]})

@app.route('/api/applicants/<int:applicant_id>', methods=['GET'])
def get_applicant(applicant_id):
    if 'user_id' not in session:
        return jsonify({'error': 'غير مصرح'}), 401
    
    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT * FROM applicants WHERE id = ?', (applicant_id,))
    applicant = cursor.fetchone()
    db.close()
    
    if not applicant:
        return jsonify({'error': 'الطالب غير موجود'}), 404
    
    return jsonify({'applicant': dict(applicant)})

@app.route('/api/applicants/<int:applicant_id>/status', methods=['PUT'])
def update_applicant_status(applicant_id):
    if 'user_id' not in session:
        return jsonify({'error': 'غير مصرح'}), 401
    
    data = request.get_json()
    status = data.get('status')
    
    if status not in ['pending', 'approved', 'rejected']:
        return jsonify({'error': 'حالة غير صالحة'}), 400
    
    db = get_db()
    cursor = db.cursor()
    cursor.execute('UPDATE applicants SET status = ? WHERE id = ?', (status, applicant_id))
    db.commit()
    db.close()
    
    return jsonify({'success': True, 'message': f'تم تحديث الحالة إلى {status}'})

@app.route('/api/applicants/export', methods=['GET'])
def export_applicants():
    if 'user_id' not in session:
        return jsonify({'error': 'غير مصرح'}), 401
    
    db = get_db()
    cursor = db.cursor()
    cursor.execute('''
        SELECT name, phone, parent_name, age, program, preferred_time, notes, submission_date, status
        FROM applicants ORDER BY submission_date DESC
    ''')
    
    applicants = cursor.fetchall()
    db.close()
    
    # إنشاء ملف Excel
    import io
    import openpyxl
    from openpyxl import Workbook
    
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "الطلاب المقدمين"
    
    # Headers
    headers = ['الاسم', 'الهاتف', 'ولي الأمر', 'العمر', 'البرنامج', 'الوقت المناسب', 'ملاحظات', 'تاريخ التقديم', 'الحالة']
    ws.append(headers)
    
    # Data
    for applicant in applicants:
        row = [
            applicant[0],  # name
            applicant[1],  # phone
            applicant[2],  # parent_name
            applicant[3],  # age
            applicant[4],  # program
            applicant[5],  # preferred_time
            applicant[6],  # notes
            applicant[7],  # submission_date
            applicant[8]   # status
        ]
        ws.append(row)
    
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name='applicants.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ==================== ADVANCED ATTENDANCE API ====================
@app.route('/api/attendance/daily', methods=['GET'])
@login_required
def get_daily_attendance():
    """الحصول على ورقة الحضور اليومية"""
    group_id = request.args.get('group_id', type=int)
    
    try:
        attendance_sheet = attendance_system.generate_daily_attendance_sheet(group_id)
        return jsonify({
            'success': True,
            'data': attendance_sheet
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/attendance/mark', methods=['POST'])
@login_required
def mark_attendance():
    """تسجيل الحضور والغياب"""
    data = request.get_json()
    attendance_data = data.get('attendance', {})
    
    try:
        result = attendance_system.mark_attendance_batch(attendance_data)
        return jsonify(result)
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/attendance/report', methods=['GET'])
@login_required
def get_attendance_report():
    """الحصول على تقرير الحضور والغياب"""
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    student_id = request.args.get('student_id', type=int)
    group_id = request.args.get('group_id', type=int)
    
    try:
        records = attendance_system.get_attendance_report(
            date_from, date_to, student_id, group_id
        )
        return jsonify({
            'success': True,
            'data': records
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/attendance/export', methods=['GET'])
@login_required
def export_attendance():
    """تصدير بيانات الحضور"""
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    group_id = request.args.get('group_id', type=int)
    
    try:
        output = attendance_system.export_attendance_excel(date_from, date_to, group_id)
        return send_file(
            output,
            as_attachment=True,
            download_name=f'attendance_report_{date_from}_to_{date_to}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ==================== AI REPORTS API ====================
@app.route('/api/reports/ai/student/<int:student_id>', methods=['GET'])
@login_required
def generate_ai_student_report(student_id):
    """توليد تقرير ذكاء اصطناعي للطالب"""
    try:
        # الحصول على بيانات الطالب
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT * FROM students WHERE id = ?', (student_id,))
        student = cursor.fetchone()
        
        if not student:
            return jsonify({'error': 'الطالب غير موجود'}), 404
        
        # الحصول على بيانات الحضور
        cursor.execute('''
            SELECT * FROM attendance 
            WHERE student_id = ? 
            ORDER BY date DESC 
            LIMIT 30
        ''', (student_id,))
        attendance_data = [dict(record) for record in cursor.fetchall()]
        
        # الحصول على بيانات الدرجات
        cursor.execute('''
            SELECT * FROM grades_results 
            WHERE student_id = ? 
            ORDER BY exam_date DESC 
            LIMIT 20
        ''', (student_id,))
        grades_data = [dict(record) for record in cursor.fetchall()]
        
        # الحصول على بيانات المدفوعات
        cursor.execute('''
            SELECT * FROM payments 
            WHERE student_id = ? 
            ORDER BY payment_date DESC 
            LIMIT 20
        ''', (student_id,))
        payments_data = [dict(record) for record in cursor.fetchall()]
        
        db.close()
        
        # توليد التقرير
        filepath, filename = generate_ai_student_report(
            dict(student), attendance_data, grades_data, payments_data
        )
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ==================== FINANCIAL MANAGEMENT API ====================
@app.route('/api/financial/summary', methods=['GET'])
@login_required
def get_financial_summary():
    """الحصول على ملخص مالي"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # إجمالي الإيرادات
        cursor.execute('SELECT SUM(amount) as total FROM payments')
        total_income = cursor.fetchone()['total'] or 0
        
        # إجمالي المصروفات
        cursor.execute('SELECT SUM(amount) as total FROM expenses')
        total_expenses = cursor.fetchone()['total'] or 0
        
        # رواتب المعلمين
        cursor.execute('SELECT SUM(amount) as total FROM teacher_salaries')
        teacher_salaries = cursor.fetchone()['total'] or 0
        
        db.close()
        
        financial_data = {
            'total_income': total_income,
            'total_expenses': total_expenses,
            'teacher_salaries': teacher_salaries,
            'period': 'شهر الحالي'
        }
        
        return jsonify({
            'success': True,
            'data': financial_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/financial/report', methods=['GET'])
@login_required
def generate_financial_report():
    """توليد التقرير المالي"""
    try:
        # الحصول على البيانات المالية
        financial_data = get_financial_summary().get_json()['data']
        
        # توليد التقرير
        filepath, filename = generate_financial_report(financial_data)
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ==================== SECURE ADMIN AREA ====================
@app.route('/api/admin/verify', methods=['POST'])
def verify_admin_access():
    """التحقق من صلاحيات المدير"""
    data = request.get_json()
    password = data.get('password')
    
    # كلمة سر خاصة للقسم المالي
    ADMIN_PASSWORD = 'massar_admin_2024'
    
    if password == ADMIN_PASSWORD:
        session['admin_verified'] = True
        return jsonify({
            'success': True,
            'message': 'تم التحقق بنجاح'
        })
    else:
        return jsonify({
            'success': False,
            'error': 'كلمة المرور غير صحيحة'
        }), 401

@app.route('/api/admin/financial', methods=['GET'])
def get_admin_financial_data():
    """الحصول على البيانات المالية للمدير فقط"""
    if not session.get('admin_verified'):
        return jsonify({'error': 'غير مصرح'}), 401
    
    try:
        # بيانات مفصلة عن الأرباح
        db = get_db()
        cursor = db.cursor()
        
        # إيرادات الشهور الماضية
        cursor.execute('''
            SELECT month, year, SUM(amount) as monthly_income
            FROM payments
            GROUP BY month, year
            ORDER BY year DESC, month DESC
            LIMIT 12
        ''')
        monthly_income = [dict(record) for record in cursor.fetchall()]
        
        # المصروفات التفصيلية
        cursor.execute('''
            SELECT category, SUM(amount) as total
            FROM expenses
            GROUP BY category
            ORDER BY total DESC
        ''')
        expenses_by_category = [dict(record) for record in cursor.fetchall()]
        
        db.close()
        
        return jsonify({
            'success': True,
            'data': {
                'monthly_income': monthly_income,
                'expenses_by_category': expenses_by_category
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

if __name__ == '__main__':
    init_db()
    print("\n" + "="*50)
    print("  مركز مسار التعليمي - نظام الإدارة المتقدم")
    print("="*50)
    print("  الرابط: http://localhost:5000")
    print("  المستخدم: admin")
    print("  كلمة المرور: admin123")
    print("  كلمة سر المدير: massar_admin_2024")
    print("="*50 + "\n")
    app.run(debug=False, host='0.0.0.0', port=5000)
